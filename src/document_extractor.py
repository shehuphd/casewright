import base64
import csv
import io
from pathlib import Path

IMAGE_EXTENSIONS = {".png", ".jpg", ".jpeg", ".webp", ".gif"}
MAX_CHARS_PER_DOC = 15000  # ~3750 tokens; keeps total prompt manageable


def extract_pdf(filepath: Path) -> dict:
    try:
        import pypdf
        reader = pypdf.PdfReader(str(filepath))
        pages = []
        for i, page in enumerate(reader.pages):
            text = (page.extract_text() or "").strip()
            if text:
                pages.append({"page": i + 1, "text": text})

        full_text = "\n\n".join(f"[Page {p['page']}]\n{p['text']}" for p in pages)
        if len(full_text) > MAX_CHARS_PER_DOC:
            full_text = full_text[:MAX_CHARS_PER_DOC] + "\n\n[... text truncated — document continues beyond this point ...]"

        return {
            "filename": filepath.name,
            "file_type": "pdf",
            "page_count": len(reader.pages),
            "pages": pages,
            "full_text": full_text,
            "status": "success" if full_text.strip() else "needs_review",
        }
    except Exception as e:
        return {
            "filename": filepath.name,
            "file_type": "pdf",
            "page_count": 0,
            "pages": [],
            "full_text": "",
            "status": "needs_review",
            "error": str(e),
        }


def extract_image(filepath: Path) -> dict:
    suffix = filepath.suffix.lower()
    try:
        from PIL import Image
        img = Image.open(filepath)
        orig_size = f"{img.width}×{img.height}"
        # Ensure we can encode it — convert palette/transparency modes
        if img.mode in ("RGBA", "P", "LA"):
            img = img.convert("RGB")
        # Resize if either dimension exceeds 2000px to keep token cost reasonable
        max_dim = 2000
        if img.width > max_dim or img.height > max_dim:
            ratio = min(max_dim / img.width, max_dim / img.height)
            img = img.resize((int(img.width * ratio), int(img.height * ratio)), Image.LANCZOS)
        buf = io.BytesIO()
        fmt = "JPEG" if suffix in (".jpg", ".jpeg") else "PNG"
        img.save(buf, format=fmt)
        b64 = base64.b64encode(buf.getvalue()).decode("utf-8")
        mime = "image/jpeg" if fmt == "JPEG" else "image/png"
    except ImportError:
        # Pillow not installed — raw encode, no resizing
        with open(filepath, "rb") as f:
            b64 = base64.b64encode(f.read()).decode("utf-8")
        mime = "image/jpeg" if suffix in (".jpg", ".jpeg") else f"image/{suffix[1:]}"
        orig_size = None
    except Exception as e:
        return {"filename": filepath.name, "file_type": "image", "status": "needs_review", "error": str(e)}

    result = {
        "filename": filepath.name,
        "file_type": "image",
        "image_data": b64,
        "image_mime": mime,
        "status": "image",
    }
    if orig_size:
        result["original_size"] = orig_size
    return result


def extract_csv(filepath: Path) -> dict:
    try:
        with open(filepath, newline="", encoding="utf-8-sig") as f:
            reader = csv.reader(f)
            rows = list(reader)
        if not rows:
            return {"filename": filepath.name, "file_type": "csv", "status": "needs_review", "full_text": ""}
        lines = []
        for i, row in enumerate(rows):
            lines.append(" | ".join(str(cell) for cell in row))
            if i == 0:
                lines.append("-" * 40)
        full_text = "\n".join(lines)
        if len(full_text) > MAX_CHARS_PER_DOC:
            full_text = full_text[:MAX_CHARS_PER_DOC] + "\n[... truncated ...]"
        return {
            "filename": filepath.name,
            "file_type": "csv",
            "row_count": len(rows),
            "full_text": full_text,
            "status": "success",
        }
    except Exception as e:
        return {"filename": filepath.name, "file_type": "csv", "status": "needs_review", "error": str(e), "full_text": ""}


def extract_xlsx(filepath: Path) -> dict:
    try:
        import openpyxl
        wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
        sheets = []
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            rows = []
            for row in ws.iter_rows(values_only=True):
                if any(cell is not None for cell in row):
                    rows.append(" | ".join("" if cell is None else str(cell) for cell in row))
            if rows:
                sheets.append(f"[Sheet: {sheet_name}]\n" + "\n".join(rows))
        full_text = "\n\n".join(sheets)
        if len(full_text) > MAX_CHARS_PER_DOC:
            full_text = full_text[:MAX_CHARS_PER_DOC] + "\n[... truncated ...]"
        return {
            "filename": filepath.name,
            "file_type": "xlsx",
            "sheet_count": len(wb.sheetnames),
            "full_text": full_text,
            "status": "success" if full_text.strip() else "needs_review",
        }
    except ImportError:
        return {"filename": filepath.name, "file_type": "xlsx", "status": "needs_review", "error": "openpyxl not installed", "full_text": ""}
    except Exception as e:
        return {"filename": filepath.name, "file_type": "xlsx", "status": "needs_review", "error": str(e), "full_text": ""}


def extract_document(filepath: Path) -> dict:
    suffix = filepath.suffix.lower()
    if suffix == ".pdf":
        return extract_pdf(filepath)
    if suffix in IMAGE_EXTENSIONS:
        return extract_image(filepath)
    if suffix == ".csv":
        return extract_csv(filepath)
    if suffix in (".xlsx", ".xls"):
        return extract_xlsx(filepath)
    return {
        "filename": filepath.name,
        "file_type": "unknown",
        "status": "needs_review",
        "full_text": "",
    }


def extract_case_documents(document_names: list[str], documents_dir: str) -> list[dict]:
    docs_path = Path(documents_dir)
    results = []
    for name in document_names:
        filepath = docs_path / name
        if not filepath.exists():
            results.append({"filename": name, "file_type": "unknown", "status": "not_found"})
        else:
            results.append(extract_document(filepath))
    return results
